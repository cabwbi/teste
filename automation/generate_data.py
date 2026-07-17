#!/usr/bin/env python3
"""Regenera exclusivamente os 16 arquivos de dados do SISCABW."""
from __future__ import annotations

import argparse
import json
import re
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

EXPECTED_SCHEMAS = {
    "NL_requisicao.xlsx": (1, {"REQUISIÇÃO", "FATURA", "PO", "VALOR PAGO NO PEDIDO", "DATA"}),
    "Ordem_de_compra_em_assinatura.xlsx": (3, {"PO", "DATA", "VALOR", "DIGITO", "ALTERAÇÃO"}),
    "controle_financeiro_contratos.xlsx": (2, {"CONTRATO", "EMPRESA", "VALOR DO CONTRATO", "GRAND COMANDO", "TOTAL FATURADO USD"}),
    "descricao_OM.xlsx": (1, {"UNID", "SIGLA", "DESCRIÇÃO"}),
    "descricao_projetos.xlsx": (1, {"PROJETO", "NONE"}),
    "digitos.xlsx": (2, {"DÍGITO", "DOTAÇÃO", "SALDO", "UNIDADES", "PROJETOS"}),
    "ordem_de_compra.xlsx": (1, {"PO", "VERSAO", "DATA", "VAL TOT USD", "TOT FATUR USD", "SALDO USD", "PAG"}),
    "requisicoes.xlsx": (1, {"PEDIDO", "PRJ", "STATUS", "NUM OC", "QTD", "QTD REC", "NOMENCLATURA", "DPE"}),
    "volumes.xlsx": (1, {"VOLUME", "PEDIDO", "PAG", "MANIFESTO"}),
}


def text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().lstrip("'")


def number(value: Any) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def money(value: Any) -> float:
    return round(number(value) + 0.0, 2)


def iso(value: Any) -> str:
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    raw = text(value)
    return raw[:10] if re.match(r"^\d{4}-\d{2}-\d{2}", raw) else ""


def workbook_rows(path: Path, header_row: int) -> tuple[list[dict[str, Any]], datetime | None]:
    wb = load_workbook(path, read_only=True, data_only=True)
    modified = wb.properties.modified
    ws = wb.active
    iterator = ws.iter_rows(values_only=True)
    for _ in range(header_row - 1):
        next(iterator)
    headers = [text(v) for v in next(iterator)]
    rows = [dict(zip(headers, row)) for row in iterator if any(v is not None for v in row)]
    return rows, modified


def split_codes(raw: Any) -> list[str]:
    value = text(raw)
    if not value:
        return []
    return [p for p in re.split(r"[,;/\s]+", value) if p]


def unique_join(values: list[str], sep: str = "; ") -> str:
    return sep.join(dict.fromkeys(v for v in values if v))


def po_year(po: Any) -> int:
    value = text(po)
    return 2000 + int(value[:2]) if len(value) >= 2 and value[:2].isdigit() else 0


def js(prefix: str, value: Any) -> str:
    return prefix + json.dumps(value, ensure_ascii=False, separators=(",", ":"), allow_nan=False) + ";\n"


def write(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--inputs", type=Path, required=True)
    parser.add_argument("--repo", type=Path, required=True)
    args = parser.parse_args()
    tables: dict[str, list[dict[str, Any]]] = {}
    modified: list[datetime] = []
    for name, (header_row, required) in EXPECTED_SCHEMAS.items():
        path = args.inputs / name
        if not path.is_file():
            raise SystemExit(f"Planilha obrigatória ausente: {name}")
        rows, stamp = workbook_rows(path, header_row)
        columns = set(rows[0]) if rows else set()
        absent = sorted(required - columns)
        if absent:
            raise SystemExit(f"Schema inválido em {name}; colunas ausentes: {absent}")
        tables[name] = rows
        if stamp:
            modified.append(stamp.replace(tzinfo=None))
    generated_at = max(modified).strftime("%Y-%m-%d %H:%M:%S") if modified else "1970-01-01 00:00:00"
    as_of = max(modified).date() if modified else date(1970, 1, 1)

    om_rows = tables["descricao_OM.xlsx"]
    om_lookup = {}
    for row in om_rows:
        code, sigla, desc = text(row.get("UNID")), text(row.get("SIGLA")), text(row.get("DESCRIÇÃO"))
        if code:
            om_lookup[code] = " - ".join(v for v in (sigla, desc) if v)
    project_lookup = {
        text(r.get("PROJETO")): text(r.get("NONE")) for r in tables["descricao_projetos.xlsx"] if text(r.get("PROJETO"))
    }
    project_label = lambda code: f"{code} - {project_lookup[code]}" if code in project_lookup else (code or "Não informado")

    requisitions = tables["requisicoes.xlsx"]
    req_by_po: dict[str, list[dict[str, Any]]] = defaultdict(list)
    req_by_id: dict[str, dict[str, Any]] = {}
    for row in requisitions:
        pedido = text(row.get("PEDIDO"))
        if pedido:
            req_by_id[pedido] = row
        po = text(row.get("NUM OC"))
        if po:
            req_by_po[po].append(row)

    orders = tables["ordem_de_compra.xlsx"]
    current_year = max((po_year(r.get("PO")) for r in orders), default=date.today().year)
    contract_rows = tables["controle_financeiro_contratos.xlsx"]
    contract_ids = {text(r.get("CONTRATO")) for r in contract_rows}

    def object_for_po(po: str) -> str:
        values = []
        for req in req_by_po.get(po, []):
            values.append(text(req.get("NOMENCLATURA")) or text(req.get("DESCRIÇÃO DO MATERIAL")))
        return unique_join(values)

    # Contratos
    contracts = []
    for row in contract_rows:
        cage, command, unit = text(row.get("CAGE")), text(row.get("GRAND COMANDO")), text(row.get("UNIDADE"))
        if cage == "W2525":
            category, category_label = "fms", "FMS (Foreign Military Sales)"
        elif command == "CW":
            category, category_label = "administrativos", "Contratos Administrativos"
        else:
            category, category_label = "finalisticos", "Contratos Finalísticos"
        end = iso(row.get("DATA FINAL"))
        if not end:
            validity, risk = "Sem data final", False
        else:
            days = (datetime.strptime(end, "%Y-%m-%d").date() - as_of).days
            if days < 0: validity, risk = "Vigência expirada", True
            elif days <= 90: validity, risk = "Vencimento em até 90 dias", False
            elif days <= 150: validity, risk = "Vencimento entre 90 e 150 dias", False
            else: validity, risk = "Vencimento acima de 150 dias", False
        ordering = "Ordenação de Despesas pela CABW" if unit == "CW" else "Ordenação de Despesas pela OM Requisitante"
        obj = text(row.get("OBJETO RESUMO"))
        contracts.append({
            "contrato": text(row.get("CONTRATO")), "numero": text(row.get("NÚMERO")), "unidade": unit,
            "cage": cage, "empresa": text(row.get("EMPRESA")), "objeto": obj, "objetoResumo": obj,
            "moeda": text(row.get("MOEDA")), "valorContrato": money(row.get("VALOR DO CONTRATO")),
            "totalEmpenhado": money(row.get("TOTAL EMPENHADO")), "totalFaturado": money(row.get("TOTAL FATURADO")),
            "saldoSiloms": money(row.get("SALDO SILOMS-EXT")), "valorAEmpenhar": money(row.get("VALOR A EMPENHAR")),
            "dataAssinatura": iso(row.get("DATA ASSINATURA")), "dataInicio": iso(row.get("DATA INÍCIO")), "dataFinal": end,
            "pressuposto": text(row.get("PRESSUPOSTO")), "grandeComando": command, "grandComando": command,
            "acao": text(row.get("AÇÃO")), "acdComp": text(row.get("ACD COMP")),
            "totalEmpenhadoUsd": money(row.get("TOTAL EMPENHADO USD")), "totalFaturadoUsd": money(row.get("TOTAL FATURADO USD")),
            "vigencia": validity, "categoria": category, "categoriaLabel": category_label, "category": category,
            "ordenador": ordering, "ordenadorDespesa": ordering, "tipoOrdenacao": ordering, "riscoVigencia": risk,
        })
    categories = ("administrativos", "finalisticos", "fms")
    summary = {
        "total": len(contracts),
        "counts": {c: sum(x["categoria"] == c for x in contracts) for c in categories},
        "totals": {k: money(sum(x[k] for x in contracts)) for k in ("valorContrato", "totalEmpenhadoUsd", "totalFaturadoUsd")},
        "byCategory": {},
    }
    for category in categories:
        selected = [x for x in contracts if x["categoria"] == category]
        summary["byCategory"][category] = {
            "count": len(selected),
            **{k: money(sum(x[k] for x in selected)) for k in ("valorContrato", "totalEmpenhadoUsd", "totalFaturadoUsd")},
        }
    po_details: dict[str, list[dict[str, Any]]] = defaultdict(list)
    po_to_pag: dict[str, str] = {}
    for row in orders:
        pag, po = text(row.get("PAG")), text(row.get("PO"))
        if pag in contract_ids:
            po_to_pag[po] = pag
            po_details[pag].append({
                "po": po, "data": iso(row.get("DATA")),
                "fornecedor": unique_join([text(row.get("FORNECEDOR")), text(row.get("NOME FORNECEDOR"))], " - "),
                "objetoResumo": object_for_po(po), "projeto": text(row.get("PROJETO")),
                "valorTotalUsd": money(row.get("VAL TOT USD")), "valorFaturadoUsd": money(row.get("TOT FATUR USD")),
                "saldoUsd": money(row.get("SALDO USD")),
            })
    monthly = {pag: {"current": [0.0] * 12, "previous": [0.0] * 12} for pag in po_details}
    monthly_details = {pag: {"current": [[] for _ in range(12)], "previous": [[] for _ in range(12)]} for pag in po_details}
    filled_invoices = 0
    for row in tables["NL_requisicao.xlsx"]:
        po, dt = text(row.get("PO")), row.get("DATA")
        pag = po_to_pag.get(po)
        if not pag or not isinstance(dt, (date, datetime)) or dt.year not in (current_year, current_year - 1):
            continue
        period = "current" if dt.year == current_year else "previous"
        value = money(row.get("VALOR PAGO NO PEDIDO"))
        monthly[pag][period][dt.month - 1] = money(monthly[pag][period][dt.month - 1] + value)
        invoice = text(row.get("FATURA"))
        if invoice: filled_invoices += 1
        monthly_details[pag][period][dt.month - 1].append({
            "po": po, "fatura": invoice, "valor": value, "data": iso(dt), "requisicao": text(row.get("REQUISIÇÃO")),
        })
    contract_data = {
        "meta": {"geradoEm": generated_at, "fonte": "controle_financeiro_contratos.xlsx atualizado", "fonteOrdemCompra": "ordem_de_compra.xlsx", "correcao": "Categorias: administrativos por Grande Comando CW; FMS por CAGE W2525; demais finalísticos. Faturas por NL_requisicao.xlsx.", "faturasDetalheTotal": filled_invoices, "faturasDetalhePreenchidas": filled_invoices},
        "summary": summary, "contracts": contracts, "records": contracts, "data": contracts,
        "poDetails": dict(po_details), "purchaseOrdersByContract": dict(po_details),
        "invoiceMonthlyByContract": monthly, "invoiceMonthlyMeta": {"currentYear": current_year, "previousYear": current_year - 1},
        "invoiceMonthlyDetailsByContract": monthly_details,
    }

    # Crédito
    digits = []
    for row in tables["digitos.xlsx"]:
        digit = text(row.get("DÍGITO"))
        if not digit:
            continue
        om_code = text(row.get("UNIDADES"))
        projects = split_codes(row.get("PROJETOS"))
        labels = [project_label(code) for code in projects] or ["Não informado"]
        om_label = om_lookup.get(om_code, om_code)
        digits.append({
            "digito": digit, "data": iso(row.get("CRIAÇÃO")), "natureza": text(row.get("NATUREZA")),
            "ugr": text(row.get("UGR")), "nomeUgr": text(row.get("NOME DA UGR")), "omUnid": om_code,
            "omSigla": om_label.split(" - ", 1)[0] if om_label else om_code, "om": om_label,
            "projetos": text(row.get("PROJETOS")), "projetoLabel": labels[0], "dotacao": money(row.get("DOTAÇÃO")),
            "saldo": money(row.get("SALDO")), "objetivo": text(row.get("OBJETIVO")), "gc": text(row.get("GRANDE COMANDO")),
            "gcDesc": text(row.get("GC DESCRIÇÃO")), "acao": text(row.get("AÇÃO")), "ptres": text(row.get("PTRES")),
            "planoInterno": text(row.get("PLANO INTERNO")), "fonte": text(row.get("FNT RECURSO")), "indicador": text(row.get("INDICADOR")),
            "omLabel": om_label, "sigla": om_label.split(" - ", 1)[0] if om_label else om_code,
            "gcDescricao": text(row.get("GC DESCRIÇÃO")), "grandComando": text(row.get("GRANDE COMANDO")), "projetosLabels": labels,
        })
    current_pos = []
    for row in orders:
        po = text(row.get("PO"))
        if po_year(po) != current_year:
            continue
        om_code, project = text(row.get("UNIDADE REQUISITANTE")), text(row.get("PROJETO"))
        om_label = om_lookup.get(om_code, om_code)
        supplier_code, supplier_name = text(row.get("FORNECEDOR")), text(row.get("NOME FORNECEDOR"))
        current_pos.append({
            "po": po, "data": iso(row.get("DATA")), "digito": text(row.get("DÍGITO")),
            "fornecedor": unique_join([supplier_code, supplier_name], " - "), "nomeFornecedor": supplier_name,
            "qtdItens": int(number(row.get("QTD ITENS"))), "projeto": project, "projetoLabel": project_label(project),
            "moeda": text(row.get("MOEDA")), "valorUsd": money(row.get("VAL TOT USD")), "faturadoUsd": money(row.get("TOT FATUR USD")),
            "saldoUsd": money(row.get("SALDO USD")), "ne": text(row.get("NE")), "pag": text(row.get("PAG")),
            "acao": text(row.get("AÇÃO")), "modalidade": text(row.get("MODALIDADE")), "natureza": text(row.get("ND")),
            "subelemento": text(row.get("SD")), "ugr": text(row.get("UGR")), "om": om_label,
            "omSigla": om_label.split(" - ", 1)[0] if om_label else om_code, "gc": "", "omLabel": om_label,
            "sigla": om_label.split(" - ", 1)[0] if om_label else om_code, "gcDescricao": "", "grandComando": "",
            "projetosLabels": [project_label(project)], "fornecedorCodigo": supplier_code,
            "objetoResumo": object_for_po(po), "pais": text(row.get("PAIS")),
        })
    signing = []
    for row in tables["Ordem_de_compra_em_assinatura.xlsx"]:
        po, supplier = text(row.get("PO")), text(row.get("FORNECEDOR"))
        signing.append({
            "po": po, "data": iso(row.get("DATA")), "digito": text(row.get("DIGITO")), "fornecedor": supplier,
            "nomeFornecedor": supplier, "valorUsd": money(row.get("VALOR")), "moeda": text(row.get("MOEDA")),
            "modalidade": text(row.get("MODALIDADE")), "om": text(row.get("OM")),
            "objetoResumo": re.sub(r"\s+", " ", text(row.get("ALTERAÇÃO"))),
        })
    credit_data = {
        "meta": {"geradoEm": generated_at, "fonte": "Atualização automática via Google Drive", "fonteOrdemCompra": "ordem_de_compra.xlsx", "purchaseOrders2026": len(current_pos), "signatureOrders": len(signing), "sourceDigits": "digitos.xlsx"},
        "digits": digits, "pos": current_pos, "signing": signing,
        "lookups": {"om": om_lookup, "projetos": project_lookup},
        "purchaseOrders": current_pos, "signatureOrders": signing,
    }

    # Suprimento de fundos
    sf_orders = []
    for row in orders:
        if text(row.get("PROJETO")) != "SF" and text(row.get("SD")) != "96":
            continue
        po, code, name, om_code = text(row.get("PO")), text(row.get("FORNECEDOR")), text(row.get("NOME FORNECEDOR")), text(row.get("UNIDADE REQUISITANTE"))
        om_label = om_lookup.get(om_code, om_code)
        sf_orders.append({
            "po": po, "data": iso(row.get("DATA")), "ano": po_year(po), "empresaCodigo": code, "empresa": name,
            "unidadeRequisitante": om_label, "empresaLabel": f"{code} - {name} / {om_label}".strip(),
            "valorTotalUsd": money(row.get("VAL TOT USD")), "valorFaturadoUsd": money(row.get("TOT FATUR USD")),
            "saldoUsd": money(row.get("SALDO USD")), "natureza": text(row.get("ND")), "subelemento": text(row.get("SD")),
            "projeto": text(row.get("PROJETO")), "pag": text(row.get("PAG")), "ne": text(row.get("NE")),
        })
    sf_digits = [{"digito": d["digito"], "saldo": d["saldo"], "acao": d["acao"], "natureza": d["natureza"], "objetivo": d["objetivo"]} for d in digits if "SF" in split_codes(d["projetos"])]
    sf_data = {
        "meta": {"geradoEm": generated_at, "fonte": "Planilhas atualizadas automaticamente"},
        "orders": sf_orders, "digits": sf_digits,
        "summary": {"totalOrders": len(sf_orders), "totalValorUsd": money(sum(x["valorTotalUsd"] for x in sf_orders)), "totalFaturadoUsd": money(sum(x["valorFaturadoUsd"] for x in sf_orders)), "totalSaldoUsd": money(sum(x["saldoUsd"] for x in sf_orders)), "totalRequisicoes": len(sf_orders)},
    }

    # Restos a pagar: saldo atual + liquidações do ano corrente.
    nl_by_po: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in tables["NL_requisicao.xlsx"]:
        dt, po = row.get("DATA"), text(row.get("PO"))
        if isinstance(dt, (date, datetime)) and dt.year == current_year and po:
            nl_by_po[po].append(row)
    rp_records = []
    for row in orders:
        po, year, raw_balance = text(row.get("PO")), po_year(row.get("PO")), number(row.get("SALDO USD"))
        # O SISCABW arredonda o saldo por PO para centavos. Saldos residuais
        # positivos abaixo de US$ 0,01 não são RP; POs com NL no ano corrente
        # permanecem para reconstruir o valor inicialmente inscrito.
        if not (2022 <= year < current_year) or money(raw_balance) < 0 or not (money(raw_balance) > 0 or po in nl_by_po):
            continue
        reqs = req_by_po.get(po, [])
        req_ids = [text(r.get("PEDIDO")) for r in reqs if text(r.get("PEDIDO"))]
        om_codes = [r[:2] for r in req_ids]
        om_labels = [om_lookup.get(c, c) for c in om_codes]
        dpe_late = []
        for req in reqs:
            dpe = req.get("DPE")
            status = text(req.get("STATUS")).upper()
            pending = number(req.get("QTD REC")) < number(req.get("QTD"))
            if isinstance(dpe, (date, datetime)) and dpe.date() < as_of and pending and "CANCEL" not in status and not status.startswith("X-"):
                dpe_late.append(iso(dpe))
        supplier_code, supplier_name = text(row.get("FORNECEDOR")), text(row.get("NOME FORNECEDOR"))
        rp_records.append({
            "po": po, "data": iso(row.get("DATA")), "anoEmpenho": year,
            "ug": om_lookup.get(text(row.get("UNIDADE REQUISITANTE")), text(row.get("UNIDADE REQUISITANTE"))),
            "ugRaw": om_lookup.get(text(row.get("UNIDADE REQUISITANTE")), text(row.get("UNIDADE REQUISITANTE"))),
            "acao": text(row.get("AÇÃO")), "natureza": text(row.get("ND")), "projeto": text(row.get("PROJETO")),
            "empresa": unique_join([supplier_code, supplier_name], " - "), "empresaCodigo": supplier_code, "pag": text(row.get("PAG")),
            "valorTotalUsd": money(row.get("VAL TOT USD")), "valorFaturadoUsd": money(row.get("TOT FATUR USD")), "saldoAtualUsd": money(row.get("SALDO USD")),
            "requisicaoAtrasada": "SIM" if dpe_late else "NÃO", "projetosReq": unique_join([text(r.get("PRJ")) for r in reqs], ","),
            "objetosResumo": object_for_po(po), "requisicoes": req_ids, "dpeAtrasadas": list(dict.fromkeys(dpe_late)),
            "tipoProcesso": "Contratos" if text(row.get("PAG")) in contract_ids else "Varejo",
            "omRequisitante": om_labels[0] if om_labels else om_lookup.get(text(row.get("UNIDADE REQUISITANTE")), text(row.get("UNIDADE REQUISITANTE"))),
            "codigosOmRequisitante": ",".join(om_codes),
        })
    rp_po_set = {r["po"] for r in rp_records}
    nl_events = []
    for po in sorted(rp_po_set):
        for row in nl_by_po.get(po, []):
            dt = row["DATA"]
            nl_events.append({"po": po, "requisicao": text(row.get("REQUISIÇÃO")), "valor": money(row.get("VALOR PAGO NO PEDIDO")), "data": iso(dt), "ano": dt.year, "mes": dt.month})
    nl_total = money(sum(x["valor"] for x in nl_events))
    for record in rp_records:
        monthly_values = [0.0] * 12
        for event in nl_events:
            if event["po"] == record["po"]:
                monthly_values[event["mes"] - 1] = money(monthly_values[event["mes"] - 1] + event["valor"])
        record["_monthly"] = monthly_values
    por_ano = {str(y): money(sum(r["saldoAtualUsd"] for r in rp_records if r["anoEmpenho"] == y)) for y in range(2022, current_year)}
    type_counts = Counter(r["tipoProcesso"] for r in rp_records)
    summary_rp = {"currentYear": current_year, "totalRegistros": len(rp_records), "totalSaldoUsd": money(sum(r["saldoAtualUsd"] for r in rp_records)), "totalSaldoRp": money(sum(r["saldoAtualUsd"] for r in rp_records)), "totalNl2026": nl_total, "porAno": por_ano, "geradoEm": generated_at, "tiposProcesso": dict(type_counts), "contratosPagFonte": "controle_financeiro_contratos.xlsx coluna CONTRATO", "omRequisitanteFonte": "descricao_OM.xlsx, por prefixo de dois caracteres da requisição vinculada à PO", "registrosOmCorrigidos": len(rp_records), "totalNlRp2026Reconstruido": nl_total}
    max_month = max((e["mes"] for e in nl_events), default=1)
    previous_month = max_month - 1 if max_month > 1 else 12
    previous_year = current_year if max_month > 1 else current_year - 1
    top_items = []
    order_by_po = {text(r.get("PO")): r for r in orders}
    for event in sorted((e for e in nl_events if e["mes"] == previous_month and e["ano"] == previous_year), key=lambda e: e["valor"], reverse=True)[:10]:
        order = order_by_po.get(event["po"], {})
        req = req_by_id.get(event["requisicao"], {})
        top_items.append({"po": event["po"], "dataPO": iso(order.get("DATA")), "empresa": unique_join([text(order.get("FORNECEDOR")), text(order.get("NOME FORNECEDOR"))], " - "), "requisicao": event["requisicao"], "descricaoRequisicao": text(req.get("NOMENCLATURA")) or text(req.get("DESCRIÇÃO DO MATERIAL")), "valorLiquidado": event["valor"], "dataLiquidacao": event["data"]})
    evolution_items = []
    for record in rp_records:
        copy = dict(record)
        copy["liquidacoes2026"] = copy.pop("_monthly")
        evolution_items.append(copy)
    clean_records = [{k: v for k, v in r.items() if k != "_monthly"} for r in rp_records]
    resumo_ano = {}
    for year in range(2022, current_year):
        actual = por_ano[str(year)]
        liquidated = money(sum(sum(r["liquidacoes2026"]) for r in evolution_items if r["anoEmpenho"] == year))
        enrolled = money(actual + liquidated)
        resumo_ano[str(year)] = {"atual": actual, "liquidado": liquidated, "inscrito": enrolled, "percentualLiquidado": round(liquidated / enrolled * 100, 2) if enrolled else 0.0}
    actual_total, enrolled_total = summary_rp["totalSaldoUsd"], money(summary_rp["totalSaldoUsd"] + nl_total)
    rp_data = {"summary": summary_rp, "records": clean_records, "nlEvents": nl_events, "topLiquidacoesMesAnterior": {"ano": previous_year, "mes": previous_month, "items": top_items}, "rpEvolution": {"currentYear": current_year, "maxMonth": max_month, "xMode": "inicio-jan-mais-meses-decorridos", "items": evolution_items, "resumoPorAno": resumo_ano, "resumoGeral": {"atual": actual_total, "liquidado": nl_total, "inscrito": enrolled_total, "percentualLiquidado": round(nl_total / enrolled_total * 100, 2) if enrolled_total else 0.0}, "totalLiquidacoes2026Usd": nl_total, "fonte": "NL_requisicao.xlsx", "criterio": "saldo inscrito reconstruído por saldo atual da PO + liquidações 2026 registradas na NL_requisicao.xlsx"}}

    outputs = {
        "contracts-data.js": js("window.CABW_CONTRACTS_DATA = ", contract_data),
        "credit-data.js": js("window.CABW_CREDIT_DATA = ", credit_data),
        "rp-data.js": js("window.CABW_RP_DATA = ", rp_data),
        "suprimento-data.js": js("window.CABW_SUPRIMENTO_DATA = ", sf_data),
        "contracts-summary.json": json.dumps(summary, ensure_ascii=False, indent=2) + "\n",
        "credit-current.json": json.dumps(credit_data, ensure_ascii=False, separators=(",", ":")) + "\n",
        "credit-10062026.json": json.dumps(credit_data, ensure_ascii=False, separators=(",", ":")) + "\n",
        "credit-11062026.json": json.dumps(credit_data, ensure_ascii=False, separators=(",", ":")) + "\n",
    }
    mirrors = {
        "contracts-data.js": "assets/js/contracts-data.js", "credit-data.js": "assets/js/credit-data.js",
        "rp-data.js": "assets/js/rp-data.js", "suprimento-data.js": "assets/js/suprimento-data.js",
        "contracts-summary.json": "assets/data/contracts-summary.json", "credit-current.json": "assets/data/credit-current.json",
        "credit-10062026.json": "assets/data/credit-10062026.json", "credit-11062026.json": "assets/data/credit-11062026.json",
    }
    for name, content in outputs.items():
        write(args.repo / name, content)
        write(args.repo / mirrors[name], content)
    print(json.dumps({"geradoEm": generated_at, "contratos": len(contracts), "pos2026": len(current_pos), "rp": len(rp_records), "sf": len(sf_orders)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
